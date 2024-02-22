import json
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from .models import CartItem, Order, OrderItem
from openpyxl.styles import Alignment, Font
from cart.models import *


@login_required
def add_to_cart(request, product_id):
    # Get the product object from the database
    product = Product.objects.get(id=product_id)

    # Get the quantity to add to the cart from the POST request
    quantity = int(request.POST.get('quantity', 1))

    # Check if there is already an item in the cart with this product
    try:
        cart_item = CartItem.objects.get(product=product, user=request.user)
        cart_item.quantity += 1  # Increase the quantity by 1
        cart_item.save()
    except CartItem.DoesNotExist:
        # If there is no cart item with this product, create a new one
        cart_item = CartItem(product=product, user=request.user)
        cart_item.save()

    # Redirect the user to the cart page
    return redirect('cart')


def cart(request):
    # Retrieve the cart items for the current user
    cart_items = CartItem.objects.filter(user=request.user)

    # Calculate the total price of all items in the cart
    total_price = sum(item.product.price * item.quantity for item in cart_items)

    context = {
        'cart_items': cart_items,
        'total_price': total_price,
    }

    return render(request, 'cart/templates/cart.html', context)


@login_required
def update_cart(request):
    if request.method == 'POST':
        for product_id in request.POST:
            try:
                # Check if the input field has a non-zero value
                quantity = int(request.POST[product_id])
                if quantity > 0:
                    # Get the product and user objects
                    product = Product.objects.get(pk=product_id)
                    user = request.user

                    # Check if the item is already in the cart
                    try:
                        cart_item = CartItem.objects.get(product=product, user=user)
                        cart_item.quantity += quantity
                        cart_item.save()
                    except CartItem.DoesNotExist:
                        CartItem.objects.create(product=product, quantity=quantity, user=user)
            except ValueError:
                pass

        messages.success(request, 'Cart updated successfully.')
        return redirect('cart')
    else:
        # Render an error message if the request method is not POST
        messages.error(request, 'Invalid request method.')
        return redirect('cart')


def remove_cart_item(request, item_id):
    cart_item = get_object_or_404(CartItem, id=item_id)
    cart_item.delete()
    messages.success(request, 'Item removed from cart!')
    return redirect('cart')


def submit_order(request):
    if request.method == 'POST':
        # Get the name and email/phone from the form
        name = request.POST['name']
        store_number = request.POST['store_number']

        # Retrieve the cart items for the current user
        cart_items = CartItem.objects.filter(user=request.user)

        # Calculate the total price of all items in the cart
        total_price = sum(item.product.price * item.quantity for item in cart_items)

        # Create a new Order object and save it to the database
        order = Order.objects.create(user=request.user, customer_name=name, storeNumber=store_number,
                                     total_price=total_price)

        # Get the items from the user's cart and add them to the order
        for cart_item in cart_items:
            # Get the quantity of the item in the cart
            quantity = cart_item.quantity

            # Get the product from the cart item
            product = cart_item.product

            # Subtract the quantity of the item from the product's stock
            product.quantity -= quantity
            product.save()

            # Create the order item and save it
            order_item = OrderItem.objects.create(order=order, product=product, quantity=quantity,
                                                  price=product.price * quantity)
            order_item.save()

        # Write order data to Excel file
        order_data = [{'product_name': item.product.vape,
                       'skew': item.product.skew, 'quantity': item.quantity, 'total_price': item.price} for item in
                      order.order_items.all()]
        write_to_excel(order_data, name, store_number, order.id)

        # Delete all cart items associated with the current user
        user_cart_items = CartItem.objects.filter(user=request.user)
        user_cart_items.delete()

        # Send a message on Discord to notify the shop owner about the new order
        webhook_url = 'https://discord.com/api/webhooks/1093787467464847400/s81zhh31-lYrqqzKWgV7oRNvcwSu0xJQpw41Dze8BFsSuLCNn0h2kHTVS-PcX0r5YJcR'
        order_items = [f'{item.quantity}   {item.product.vape}   (${item.price:.2f})' for item in
                       order.order_items.all()]
        order_list = '\n'.join(order_items)
        message = f'**New Order #{order.id}**\n\nName: {name}\nStore: {store_number}\n\n**Order Items**:\n{order_list}\n\nTotal Price: ${total_price:.2f}'
        payload = {'content': message, 'allowed_mentions': {'parse': []}}
        headers = {'Content-Type': 'application/json'}
        response = requests.post(webhook_url, data=json.dumps(payload), headers=headers)
        if response.status_code != 204:
            print(f'Error sending message to Discord: {response.text}')

        # Redirect the user to the shop page and show a success message
        messages.info(request, 'Your order has been placed!')
        return redirect('shop')
    else:
        return redirect('cart')


def write_to_excel(order_data, name, store_number, order_id):
    # Create a new Workbook
    wb = Workbook()

    # Select the active sheet
    sheet = wb.active

    # Add header for the document
    sheet.merge_cells('A1:D1')
    sheet['A1'] = 'RPF Vape Transfer Sheet'
    sheet['A2'] = f'Orderer:'
    sheet['A3'] = f'Store Number:'
    sheet['A4'] = f'Order Number:'
    sheet['B2'] = name
    sheet['B3'] = store_number
    sheet['B4'] = order_id

    # Skip rows 4 and 5 for the gap
    sheet['A5'] = ''  # add a blank row for spacing
    sheet['A6'] = 'Product Name:'
    sheet['B6'] = 'Skew:'
    sheet['C6'] = 'Quantity:'
    sheet['D6'] = 'Price:'

    # Set column widths
    column_widths = {'A': 25, 'B': 15, 'C': 10, 'D': 10}
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    # Apply styling to headers
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    for col in ['A', 'B', 'C', 'D', 'E']:
        sheet[f'{col}5'].font = header_font
        sheet[f'{col}5'].alignment = header_alignment

    # Write order data to the sheet
    for row_num, order in enumerate(order_data, 7):
        sheet.cell(row=row_num, column=1, value=order['product_name'])
        sheet.cell(row=row_num, column=2, value=order.get('skew', ''))
        sheet.cell(row=row_num, column=3, value=order['quantity'])
        sheet.cell(row=row_num, column=4, value=order['total_price'])

    # Save the workbook to a file
    wb.save('order_sheet.xlsx')
